import cv2
from pyzbar.pyzbar import decode
import numpy as np
import serial
import time
import os
import qrcode
from PIL import Image
import math
import time 
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# esp32_port = '/dev/ttyUSB0'  # Cambia esto al puerto correcto de tu ESP32
# baud_rate = 115200  # Asegúrate de que coincida con la configuración del ESP32  
# time.sleep(2)  # Espera a que el puerto se inicialice
# try:
#     ser = serial.Serial(esp32_port, baud_rate, timeout=1)
#     print(f"Conectado a {esp32_port} a {baud_rate} baudios.")
# except serial.SerialException as e:
#     print(f"Error al conectar al puerto {esp32_port}: {e}")
#     exit()

nombre_carnet = 'nombre_carnet.xlsx'

if not os.path.exists("Asistencias"):
    os.makedirs("Asistencias")
if not os.path.exists(os.path.join("Asistencias", "QRs")):
    os.makedirs(os.path.join("Asistencias", "QRs")) 

archivo_asistencias = 'Asistencias/asistencia.xlsx'
columnas_asitencias = ["Nombre", "Carnet", "Fecha", "Hora"]

def generar_chivos(archivo, columnas):
    if not os.path.exists(archivo):
        print("El Archivo No Existe, Creando Uno Nuevo...")
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos"
        for idx, columna in enumerate(columnas, start=1):
            letra = get_column_letter(idx)
            ws[f"{letra}1"] = columna
        wb.save(archivo)
    else:
        print("El Archivo ya existe, Agregando Datos..")
        wb = load_workbook(archivo)
        ws = wb.active

generar_chivos(archivo_asistencias, columnas_asitencias)



def verificar_qr(dato_leido):
    wb = Workbook
    wb = load_workbook(nombre_carnet)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        if row[1] == dato_leido:
            print("Carnet Correcto")
            return True
        else:
            print("Carnet Invalido")


def leer_qr_camara():
    cap = cv2.VideoCapture(0)  # Usar cámara 0 (predeterminada)

    print("Escanea el código QR con la cámara... Presiona 'q' para salir.")

    while True:
        ret, frame = cap.read()
        if not ret:
            print("No se pudo acceder a la cámara.")
            break

        # Decodificar los códigos QR en la imagen
        codigos = decode(frame)
        if codigos:
            for codigo in codigos:
                data = codigo.data.decode('utf-8')
                all_data = codigo.data.decode('utf-8', errors='ignore')
                print(all_data)
                if "Carnet:" in all_data and "Nombre:" in all_data:
                    carnet = all_data.split("Carnet:")[1].split("Nombre:")[0].strip()
                    print(carnet)
                else:
                    print("Formato incorrecto: faltan 'Carnet:' o 'Nombre:' en la cadena.")
                print(carnet)
                nombre = all_data.split("Nombre:")[1].strip()
                print(nombre)
                #
                verificar_qr(carnet)
                if verificar_qr(carnet):    
                    print("Código QR detectado:", data)
                    # ser.write(b'1')  # Enviar señal al ESP32 si es válido
                    # Guardar asistencia en el archivo de Excel
                    wb = load_workbook(archivo_asistencias)
                    ws = wb.active
                    hora_actual = time.strftime("%H:%M:%S")
                    fecha_actual = time.strftime("%Y-%m-%d")
                    nueva_fila = [nombre, carnet, fecha_actual, hora_actual]
                    ya_registrado = False
                    for row in ws.iter_rows(min_row=3, values_only=True):
                        if row[0] == nombre and row[1] == carnet and row[2] == fecha_actual:
                            print("La asistencia ya fue registrada hoy. No se guardará de nuevo.")
                            ya_registrado = True
                            break

                    if not ya_registrado:
                        hora_actual = time.strftime("%H:%M:%S")
                        nueva_fila = [nombre, carnet, fecha_actual, hora_actual]
                        print("Guardando asistencia...")
                        print(nueva_fila)
                        ws.append(nueva_fila)
                        wb.save(archivo_asistencias)
                else:
                    print("Código QR no válido, no se guardará la asistencia.")
                    # ser.write(b'0')  # Enviar señal al ESP32 si es inválido

                # Dibujar un rectángulo y mostrar el contenido
                puntos = codigo.polygon
                pts = [(p.x, p.y) for p in puntos]
                cv2.polylines(frame, [np.array(pts, np.int32)], True, (0,255,0), 3)
                x, y, w, h = codigo.rect
                cv2.putText(frame, data, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255,0,0), 2)
        else:
            print("No se detectó ningún código QR.")
            # ser.write(b'0')
        # Mostrar el video
        cv2.imshow("Lector de QR", frame)

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()
    # ser.close()

# Llamar la función
leer_qr_camara()
