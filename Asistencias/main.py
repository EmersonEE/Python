import os
import cv2
import numpy as np
import time
from pyzbar.pyzbar import decode
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import qrcode
from PIL import Image
import math



# Configuración de archivos y directorios
NOMBRE_CARNET = 'nombre_carnet.xlsx'
DIR_ASISTENCIAS = 'Asistencias'
ARCHIVO_USUARIOS = 'usuarios_creados.xlsx'
DIR_QRS = os.path.join(DIR_ASISTENCIAS, "QRs")
ARCHIVO_ASISTENCIAS = os.path.join(DIR_ASISTENCIAS, 'asistencia.xlsx')
COLUMNAS_ASISTENCIAS = ["Nombre", "Carnet", "Fecha", "Hora"]
COLUMNAS_USUARIO= ["Nombre","Carnet" "Edad","Telefono","Encargado","Telefono Encargado","Correo Encargado"]
COLUMNAS_NC = ["Carnet", "Nombre"]

def solicitar_edad():
    while True:
        try:
            numero = int(input("Ingrese Su Edad "))
            if numero > 0:
                return numero      
            else:
                print("El número debe ser positivo.")
        except ValueError:
            print("Entrada inválida. Por favor, ingrese un número entero.")

def solicitar_carnet():
    while True:
        try:
            numero = int(input("Ingrese Su Carnet "))
            if numero > 0:
                if len(str(numero)) == 9:
                    return numero
                else:
                    print("El número debe tener exactamente 9 dígitos.")
            else:
                print("El número debe ser positivo.")
        except ValueError:
            print("Entrada inválida. Por favor, ingrese un número entero.")

def solicitar_numero_telefono():
    while True:
        try:
            numero = int(input("Ingrese Su Numero de Teléfono: "))
            if numero > 0:
                if len(str(numero)) == 8:
                    return numero
                else:
                    print("El número debe tener exactamente 8 dígitos.")
            else:
                print("El número debe ser positivo.")
        except ValueError:
            print("Entrada inválida. Por favor, ingrese un número entero.")

def solicitar_string():
    while True:
        texto = input("Ingrese Su Nombre: ")
        if texto.strip():
            if any(char.isdigit() for char in texto):
                print("El nombre no debe contener números.")
            else:
                return texto
        else:
            print("El texto no puede estar vacío.")

def solicitar_string_carrera():
    while True:
        texto = input("Ingrese Su Carrera: ")
        if texto.strip():
            if any(char.isdigit() for char in texto):
                print("El nombre no debe contener números.")
            else:
                return texto
        else:
            print("El texto no puede estar vacío.")

def solicitar_string_pariente():
    while True:
        texto = input("Ingrese Nombre del Encargado: ")
        if texto.strip():
            if any(char.isdigit() for char in texto):
                print("El nombre no debe contener números.")
            else:
                return texto
        else:
            print("El texto no puede estar vacío.")

def generar_chivos(archivo, columnas):
    if not os.path.exists(archivo):
        print("El Archivo No Existe, Creando Uno Nuevo...")
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos"
        for idx, columna in enumerate(columnas, start=1):
            letra = get_column_letter(idx)
            ws[f"{letra}1"] = columna
        wb.save(archivo)
    else:
        print("El Archivo ya existe, Agregando Datos..")
        wb = load_workbook(archivo)
        ws = wb.active

def obtener_siguiente_carnet(NOMBRE_CARNET):
    try:
        wb = load_workbook(NOMBRE_CARNET)
        ws = wb.active
        carnets = []
        for row in ws.iter_rows(min_row=2, values_only=True):  
            if row[1] is not None:
                carnets.append(int(row[1]))
        if carnets:
            return str(max(carnets) + 1)
        else:
            return "202500000"
    except FileNotFoundError:
        return "202500000"


def agregar_datos_nqr(nombre,carnet):
        generar_chivos(NOMBRE_CARNET,COLUMNAS_NC)
        wb = load_workbook(NOMBRE_CARNET)
        ws = wb.active
        max_row = ws.max_row
        nueva_fila = [nombre,carnet]
        ws.append(nueva_fila)
        wb.save(NOMBRE_CARNET)

def crear_qr(nombre, carnet):
    qr_data = f"Carnet: {carnet}\nNombre: {nombre}"
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(qr_data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    ruta = os.path.join(DIR_QRS, f"{carnet}.png")
    if not os.path.exists(DIR_QRS):
        os.makedirs(DIR_QRS)
    img.save(ruta)  

def agregar_datos_usuario(nombre,carnet,edad,telefono,nombre_encargado,telefono_encargado,correo_encargado):
    generar_chivos(ARCHIVO_USUARIOS,COLUMNAS_USUARIO)
    wb = load_workbook(ARCHIVO_USUARIOS)
    ws = wb.active
    nueva_fila = [nombre,carnet, edad, telefono, nombre_encargado,telefono_encargado,correo_encargado]
    for row in ws.iter_rows(min_row=6,values_only=True):
        if row[0] == nombre and row[1] == carnet and row[2] == edad and row[3] == telefono and row[4] == nombre_encargado and row[5] == telefono_encargado and row[6] == correo_encargado:
            print("El Usuario ya existe")
            return
    ws.append(nueva_fila)
    wb.save(ARCHIVO_USUARIOS)
    agregar_datos_nqr(nombre,carnet)
    crear_qr(nombre, carnet)
    print("Usuario Agregado")


def inicializar_directorios():
    """Crea los directorios necesarios si no existen"""
    os.makedirs(DIR_QRS, exist_ok=True)

def inicializar_archivo_asistencias():
    """Crea el archivo de asistencias con las columnas necesarias si no existe"""
    if not os.path.exists(ARCHIVO_ASISTENCIAS):
        print("El Archivo No Existe, Creando Uno Nuevo...")
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos"
        for idx, columna in enumerate(COLUMNAS_ASISTENCIAS, start=1):
            ws[f"{get_column_letter(idx)}1"] = columna
        wb.save(ARCHIVO_ASISTENCIAS)
    else:
        print("El Archivo ya existe, listo para agregar datos.")

def verificar_carnet(carnet: str) -> bool:
    """Verifica si el carnet existe en la base de datos"""
    try:
        wb = load_workbook(NOMBRE_CARNET)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            if row[1] == carnet:
                print("Carnet Correcto")
                return True
        print("Carnet Inválido")
        return False
    except Exception as e:
        print(f"Error al verificar carnet: {e}")
        return False

def procesar_datos_qr(data: str) -> tuple:
    """Extrae y valida los datos del QR"""
    try:
        if "Carnet:" not in data or "Nombre:" not in data:
            print("Formato incorrecto: faltan 'Carnet:' o 'Nombre:' en la cadena.")
            return None, None
        
        carnet = data.split("Carnet:")[1].split("Nombre:")[0].strip()
        nombre = data.split("Nombre:")[1].strip()
        return nombre, carnet
    except Exception as e:
        print(f"Error al procesar QR: {e}")
        return None, None

def registrar_asistencia(nombre: str, carnet: str) -> bool:
    """Registra la asistencia en el archivo Excel si no está ya registrada"""
    try:
        wb = load_workbook(ARCHIVO_ASISTENCIAS)
        ws = wb.active
        
        fecha_actual = time.strftime("%Y-%m-%d")
        hora_actual = time.strftime("%H:%M:%S")
        
        # Verificar si ya está registrado hoy
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] == carnet and row[2] == fecha_actual:
                print("La asistencia ya fue registrada hoy.")
                return False
        
        # Registrar nueva asistencia
        nueva_fila = [nombre, carnet, fecha_actual, hora_actual]
        ws.append(nueva_fila)
        wb.save(ARCHIVO_ASISTENCIAS)
        print("Asistencia registrada exitosamente.")
        return True
    except Exception as e:
        print(f"Error al registrar asistencia: {e}")
        return False

def mostrar_frame(frame, codigo=None, data=None):
    """Muestra el frame con anotaciones opcionales para el código QR"""
    if codigo:
        # Dibujar polígono alrededor del QR
        puntos = codigo.polygon
        pts = [(p.x, p.y) for p in puntos]
        cv2.polylines(frame, [np.array(pts, np.int32)], True, (0, 255, 0), 3)
        
        # Mostrar datos del QR
        x, y, w, h = codigo.rect
        cv2.putText(frame, data, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 0, 0), 2)
    
    cv2.imshow("Lector de QR", frame)

def leer_qr_camara():
    """Función principal para leer códigos QR desde la cámara"""
    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        print("No se pudo acceder a la cámara.")
        return

    print("Escanea el código QR con la cámara... Presiona 'q' para salir.")

    while True:
        ret, frame = cap.read()
        if not ret:
            print("Error al capturar frame.")
            break

        # Decodificar códigos QR
        codigos = decode(frame)
        if codigos:
            for codigo in codigos:
                try:
                    data = codigo.data.decode('utf-8')
                    nombre, carnet = procesar_datos_qr(data)
                    
                    if nombre and carnet and verificar_carnet(carnet):
                        registrar_asistencia(nombre, carnet)
                        # ser.write(b'1')  # Enviar señal al ESP32 si es válido
                    else:
                        print("Código QR no válido.")
                        # ser.write(b'0')  # Enviar señal al ESP32 si es inválido
                    
                    mostrar_frame(frame, codigo, data)
                except Exception as e:
                    print(f"Error al procesar código QR: {e}")
                    mostrar_frame(frame)
        else:
            mostrar_frame(frame)

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()
    # ser.close()
if __name__ == "__main__":
    numero = int(input("Ingrese 1 para registrar usuario o 2 para tomar asistencia: "))
    if numero == 1:
        print("Bienvenido al Registro de Usuarios")
        nombre = solicitar_string()
        edad = solicitar_edad()
        carnet = obtener_siguiente_carnet(NOMBRE_CARNET)
        telefono = solicitar_numero_telefono()
        encargado = solicitar_string_pariente()
        telefono_encargado = solicitar_numero_telefono()
        correo_env = input("Ingrese un Correo: ")
        agregar_datos_usuario(nombre,carnet, edad, telefono, encargado, telefono_encargado, correo_env)
    elif numero == 2:
        print("Tomar Asistencia")
        inicializar_directorios()
        inicializar_archivo_asistencias()
        leer_qr_camara()