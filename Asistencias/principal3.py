import os
import cv2
import numpy as np
import time
from pyzbar.pyzbar import decode
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import qrcode
from PIL import Image, ImageTk, ImageDraw, ImageFont
import math
import os
from dotenv import load_dotenv
from email.message import EmailMessage
import ssl
import smtplib
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import paho.mqtt.client as mqtt


BROKER = "192.168.1.136"
PORT = 1883
TOPIC = "/leds"
TOPIC_PANTALLA = "/pantalla"

load_dotenv()

# Configuración de correo electrónico
password = os.getenv("PASSWORD")
email_sender = os.getenv("EMAIL")


NOMBRE_CARNET = 'nombre_carnet.xlsx'
DIR_ASISTENCIAS = 'Asistencias'
ARCHIVO_USUARIOS = 'usuarios_creados.xlsx'
DIR_QRS = os.path.join(DIR_ASISTENCIAS, "QRs")
ARCHIVO_ASISTENCIAS = os.path.join(DIR_ASISTENCIAS, 'asistencia.xlsx')
ARCHIVO_ESTADISTICAS = os.path.join(DIR_ASISTENCIAS, 'estadisticas.xlsx')
COLUMNAS_ASISTENCIAS = ["Nombre", "Carnet", "Fecha", "Hora"]
# COLUMNAS_USUARIO= ["Nombre","Carnet" "Edad","Telefono","Encargado","Telefono Encargado","Correo Encargado"]
COLUMNAS_USUARIO = [
    "Nombre", "Carnet", "Edad", "Telefono", 
    "Encargado", "Telefono Encargado", "Correo Encargado", 
    "Carrera", "Estado"  # Agregamos Carrera y Estado
]
COLUMNAS_NC = ["Carnet", "Nombre"]
class SistemaAsistenciasApp:

    def __init__(self, root):
        self.root = root
        self.root.title("Sistema de Gestión de Asistencias")
        self.root.geometry("1100x750")  # Tamaño ligeramente mayor
        self.root.minsize(1000, 700)  # Tamaño mínimo
        self.root.configure(bg='#f8f9fa')  # Fondo coherente con el estilo
        
        # Configuración inicial de estilos
        self.aplicar_estilos()
        
        # Conexión MQTT
        self.mqtt_cliente = mqtt.Client()
        self.mqtt_cliente.connect(BROKER, PORT, 60)
        self.mqtt_cliente.loop_start()
        
        # Variables para la cámara
        self.cap = None
        self.running = False
        
        # Crear contenedor principal con sombra visual
        main_container = ttk.Frame(self.root, style='Card.TFrame')
        main_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Crear el notebook (pestañas) con estilo mejorado
        self.notebook = ttk.Notebook(main_container)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Crear las pestañas
        self.crear_pestana_registro()
        self.crear_pestana_asistencia()
        self.crear_pestana_reportes()
        self.crear_pestana_verificacion()
        
        # Inicialización de archivos y directorios
        self.inicializar_directorios()
        self.inicializar_archivo_asistencias()
        # self.inicializar_archivo_estadisticas()
        
        # Añadir un footer con información
        self.crear_footer()

    def crear_footer(self):
        footer = ttk.Frame(self.root, style='Footer.TFrame')
        footer.pack(fill=tk.X, pady=(5, 0))
        
        version_label = ttk.Label(
            footer, 
            text="Sistema de Gestión de Asistencias v1.0", 
            style='Footer.TLabel'
        )
        version_label.pack(side=tk.LEFT, padx=10)
        
        status_label = ttk.Label(
            footer, 
            text="Conectado", 
            style='Status.TLabel'
        )
        status_label.pack(side=tk.RIGHT, padx=10)

    def aplicar_estilos(self):
        style = ttk.Style()
        
        # Usar tema moderno con elementos planos
        style.theme_use('clam')
        
        # Paleta de colores moderna y profesional
        colors = {
            'primary': '#2563eb',  # Azul vibrante
            'primary_dark': '#1e40af',
            'secondary': '#7c3aed',  # Púrpura
            'accent': '#06b6d4',  # Cyan
            'light_bg': '#f8fafc',
            'lighter_bg': '#ffffff',
            'dark_text': '#1e293b',
            'medium_text': '#475569',
            'light_text': '#f8fafc',
            'success': '#10b981',
            'warning': '#f59e0b',
            'danger': '#ef4444',
            'border': '#e2e8f0'
        }
        
        # Configuración general
        style.configure('.', font=('Segoe UI', 10))
        
        # Estilo para tarjetas/contenedores
        style.configure('Card.TFrame', 
                    background=colors['lighter_bg'],
                    borderwidth=1,
                    relief='solid',
                    bordercolor=colors['border'])
        
        # Estilo para footer
        style.configure('Footer.TFrame', background=colors['primary'])
        style.configure('Footer.TLabel', 
                    background=colors['primary'],
                    foreground=colors['light_text'],
                    font=('Segoe UI', 9))
        style.configure('Status.TLabel', 
                    background=colors['primary'],
                    foreground='#86efac',  # Verde claro
                    font=('Segoe UI', 9, 'bold'))
        
        # Configuración de etiquetas
        style.configure('TLabel', 
                    background=colors['light_bg'], 
                    font=('Segoe UI', 10), 
                    foreground=colors['dark_text'])
        
        style.configure('Header.TLabel', 
                    background=colors['primary'], 
                    foreground=colors['light_text'], 
                    font=('Segoe UI Semibold', 16),
                    padding=10)
        
        style.configure('Subheader.TLabel',
                    font=('Segoe UI Semibold', 12),
                    foreground=colors['primary_dark'])
        
        # Botones modernos
        style.configure('TButton', 
                    font=('Segoe UI Semibold', 10), 
                    background=colors['primary'], 
                    foreground=colors['light_text'],
                    borderwidth=0,
                    padding=(12, 6),
                    focuscolor=colors['primary'] + '00')  # Transparente
                    
        style.map('TButton', 
                background=[('active', colors['primary_dark']), 
                            ('pressed', colors['primary_dark'])], 
                foreground=[('active', colors['light_text']), 
                            ('pressed', colors['light_text'])])
        
        # Botones de acción
        style.configure('Accent.TButton',
                    background=colors['accent'],
                    foreground=colors['light_text'])
        style.map('Accent.TButton',
                background=[('active', '#0891b2'), 
                            ('pressed', '#0e7490')])
        
        # Treeview (tablas modernas)
        style.configure('Treeview.Heading', 
                    font=('Segoe UI Semibold', 10), 
                    background=colors['primary'], 
                    foreground=colors['light_text'],
                    padding=5,
                    relief='flat')
                    
        style.configure('Treeview', 
                    background=colors['lighter_bg'], 
                    foreground=colors['dark_text'], 
                    fieldbackground=colors['lighter_bg'],
                    rowheight=28,
                    bordercolor=colors['border'],
                    lightcolor=colors['lighter_bg'],
                    darkcolor=colors['lighter_bg'])
                    
        style.map('Treeview', 
                background=[('selected', colors['primary'] + '80')],  # 50% transparente
                foreground=[('selected', colors['dark_text'])])
        
        # Pestañas estilo moderno
        style.configure("TNotebook", 
                    background=colors['light_bg'],
                    tabmargins=(0, 5, 0, 0))
                    
        style.configure("TNotebook.Tab", 
                    font=('Segoe UI Semibold', 10), 
                    padding=[15, 6],
                    background=colors['light_bg'],
                    foreground=colors['medium_text'],
                    borderwidth=0,
                    focuscolor=colors['light_bg'])
                    
        style.map("TNotebook.Tab", 
                background=[("selected", colors['lighter_bg'])], 
                foreground=[("selected", colors['primary'])],
                expand=[("selected", [1, 1, 1, 0])])
        
        # Barras de desplazamiento modernas
        style.configure("Vertical.TScrollbar", 
                    arrowsize=12,
                    gripcount=0,
                    background=colors['primary'] + '40',  # 25% transparente
                    troughcolor=colors['light_bg'],
                    bordercolor=colors['light_bg'],
                    arrowcolor=colors['primary'],
                    lightcolor=colors['primary'],
                    darkcolor=colors['primary'])
        
        # Campos de entrada modernos
        style.configure('TEntry', 
                    fieldbackground=colors['lighter_bg'],
                    foreground=colors['dark_text'],
                    bordercolor=colors['border'],
                    lightcolor=colors['border'],
                    darkcolor=colors['border'],
                    padding=5,
                    insertcolor=colors['accent'])
                    
        style.map('TEntry',
                fieldbackground=[('readonly', colors['light_bg'])],
                foreground=[('readonly', colors['medium_text'])],
                bordercolor=[('focus', colors['accent'])])
        
        # Combobox moderno
        style.configure('TCombobox',
                    fieldbackground=colors['lighter_bg'],
                    foreground=colors['dark_text'],
                    selectbackground=colors['primary'],
                    selectforeground=colors['light_text'],
                    padding=5)
        style.map('TCombobox',
                fieldbackground=[('readonly', colors['lighter_bg'])],
                foreground=[('readonly', colors['dark_text'])],
                bordercolor=[('focus', colors['accent'])])
        

    def generar_estadisticas_usuario(self):
        """Genera un archivo Excel con estadísticas de asistencia por usuario"""
        try:
            # Cargar usuarios y asistencias
            wb_usuarios = load_workbook(ARCHIVO_USUARIOS)
            ws_usuarios = wb_usuarios.active

            wb_asistencias = load_workbook(ARCHIVO_ASISTENCIAS)
            ws_asistencias = wb_asistencias.active

            # Organizar asistencias por carnet
            asistencias_por_carnet = {}
            for row in ws_asistencias.iter_rows(min_row=2, values_only=True):
                nombre, carnet, fecha, hora = row
                if carnet not in asistencias_por_carnet:
                    asistencias_por_carnet[carnet] = []
                asistencias_por_carnet[carnet].append(fecha)

            # Crear directorio si no existe
            reportes_dir = os.path.join(DIR_ASISTENCIAS, "Reportes_Estadisticas")
            os.makedirs(reportes_dir, exist_ok=True)

            # Nombre del archivo
            nombre_archivo = os.path.join(reportes_dir, "estadisticas_asistencia.xlsx")

            # Crear libro de Excel
            wb = Workbook()
            ws_resumen = wb.active
            ws_resumen.title = "Resumen General"

            # Encabezados
            encabezados = [
                "Nombre", "Carnet", "Total Registros", 
                "Asistencias", "Ausencias", "Porcentaje Asistencia"
            ]
            for col_num, data in enumerate(encabezados, 1):
                ws_resumen.cell(row=1, column=col_num, value=data)

            fila = 2
            detalles = []

            for row in ws_usuarios.iter_rows(min_row=2, values_only=True):
                nombre = row[0]
                carnet = str(row[1])

                todas_fechas = sorted(set([fecha for _, _, fecha, _ in ws_asistencias.iter_rows(min_row=2, values_only=True)]))
                total_dias = len(todas_fechas)

                asistencias_usuario = asistencias_por_carnet.get(carnet, [])
                asistencias_count = len(asistencias_usuario)
                ausencias_count = total_dias - asistencias_count
                porcentaje = (asistencias_count / total_dias * 100) if total_dias > 0 else 0

                # Agregar a resumen general
                ws_resumen.append([
                    nombre, carnet, total_dias,
                    asistencias_count, ausencias_count, f"{porcentaje:.2f}%"
                ])

                # Detalles por usuario para otra hoja
                fechas_presentes = set(asistencias_usuario)
                fechas_ausentes = [f for f in todas_fechas if f not in fechas_presentes]
                detalles.append({
                    "nombre": nombre,
                    "carnet": carnet,
                    "fechas_ausentes": fechas_ausentes
                })

            # Hoja de detalles
            ws_detalles = wb.create_sheet(title="Detalles Ausencias")
            ws_detalles.append(["Nombre", "Carnet", "Fechas Ausentes"])

            for detalle in detalles:
                fechas_str = ", ".join(detalle["fechas_ausentes"]) if detalle["fechas_ausentes"] else "Ninguna"
                ws_detalles.append([detalle["nombre"], detalle["carnet"], fechas_str])

            # Ajustar anchos de columna
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    ws.column_dimensions[column].width = adjusted_width

            # Guardar archivo
            wb.save(nombre_archivo)
            messagebox.showinfo("Éxito", f"Estadísticas generadas correctamente en:\n{nombre_archivo}")

        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al generar las estadísticas: {str(e)}")
        
    
    def inicializar_directorios(self):
        os.makedirs(DIR_QRS, exist_ok=True)
    
    def generar_tarjeta(self, nombre, carrera, telefono, carnet):
        """Genera una tarjeta con los datos del usuario y su QR"""
        try:
            # Crea la carpeta si no existe
            os.makedirs("Tarjetas", exist_ok=True)
            
            # Abre la imagen base
            imagen = Image.open("1.png").convert("RGBA")
            dibujo = ImageDraw.Draw(imagen)
            
            try:
                # fuente = ImageFont.truetype("arialbd.ttf", 38)  # Windows
                fuente = ImageFont.truetype("/usr/share/fonts/TTF/DejaVuSans-Bold.ttf", 38)  # Linux
            except:
                fuente = ImageFont.load_default()
            
            # Función para texto con sombra
            def texto_con_sombra(draw, posicion, texto, fuente, color_texto="white", color_sombra="black", desplazamiento=(2, 2)):
                x, y = posicion
                dx, dy = desplazamiento
                draw.text((x + dx, y + dy), texto, font=fuente, fill=color_sombra)
                draw.text((x, y), texto, font=fuente, fill=color_texto)
            
            # Agrega los textos
            texto_con_sombra(dibujo, (50, 100), nombre, fuente)
            texto_con_sombra(dibujo, (125, 275), carrera, fuente)
            texto_con_sombra(dibujo, (125, 360), telefono, fuente)
            texto_con_sombra(dibujo, (125, 450), carnet, fuente)
            
            # Agrega el QR (que ya generaste en crear_qr)
            qr_path = os.path.join(DIR_QRS, f"{carnet}.png")
            if os.path.exists(qr_path):
                logo = Image.open(qr_path).convert("RGBA")
                logo = logo.resize((325, 325))
                imagen.paste(logo, (725, 50), logo)
            
            # Guarda la tarjeta
            tarjeta_path = os.path.join("Tarjetas", f"tarjeta_{carnet}.png")
            imagen.save(tarjeta_path)
            print(f"Tarjeta generada: {tarjeta_path}")
            
        except Exception as e:
            print(f"Error al generar tarjeta: {e}")

    def generar_chivos(self, archivo, columnas):
        if not os.path.exists(archivo):
            print("El Archivo No Existe, Creando Uno Nuevo...")
            wb = Workbook()
            ws = wb.active
            ws.title = "Datos"
            for idx, columna in enumerate(columnas, start=1):
                letra = get_column_letter(idx)
                ws[f"{letra}1"] = columna
            wb.save(archivo)
        else:
            print("El Archivo ya existe, Agregando Datos..")
            wb = load_workbook(archivo)
            ws = wb.active

    def obtener_siguiente_carnet(self,NOMBRE_CARNET):
        try:
            wb = load_workbook(NOMBRE_CARNET)
            ws = wb.active
            carnets = []
            for row in ws.iter_rows(min_row=2, values_only=True):  
                if row[1] is not None:
                    carnets.append(int(row[1]))
            if carnets:
                return str(max(carnets) + 1)
            else:
                return "202500000"
        except FileNotFoundError:
            return "202500000"

    def agregar_datos_nqr(self, nombre, carnet):
        self.generar_chivos(NOMBRE_CARNET, COLUMNAS_NC)
        wb = load_workbook(NOMBRE_CARNET)
        ws = wb.active
        nueva_fila = [nombre, carnet]
        ws.append(nueva_fila)
        wb.save(NOMBRE_CARNET)
    

    def crear_qr(self,nombre, carnet):
        qr_data = f"Carnet: {carnet}\nNombre: {nombre}"
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(qr_data)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        ruta = os.path.join(DIR_QRS, f"{carnet}.png")
        if not os.path.exists(DIR_QRS):
            os.makedirs(DIR_QRS)
        img.save(ruta)  

    def agregar_datos_usuario(self, nombre, carnet, edad, telefono, nombre_encargado, telefono_encargado, correo_encargado,carrera):
        self.generar_chivos(ARCHIVO_USUARIOS, COLUMNAS_USUARIO)
        wb = load_workbook(ARCHIVO_USUARIOS)
        ws = wb.active
        nueva_fila = [nombre, carnet, edad, telefono, nombre_encargado, telefono_encargado, correo_encargado,carrera]
        for row in ws.iter_rows(min_row=7, values_only=True):
            if row[0] == nombre and row[1] == carnet and row[2] == edad and row[3] == telefono and row[4] == nombre_encargado and row[5] == telefono_encargado and row[6] == correo_encargado:
                print("El Usuario ya existe")
                return
        ws.append(nueva_fila)
        wb.save(ARCHIVO_USUARIOS)
        self.agregar_datos_nqr(nombre, carnet)
        self.crear_qr(nombre, carnet)
        partes_nombre = nombre.split()
        
        if len(partes_nombre)> 2:
            new_name = partes_nombre[0] + " " + partes_nombre[-2]
        else:
            new_name = nombre 
        self.generar_tarjeta(new_name, carrera, telefono, carnet)
        print("Usuario Agregado")
            
    def inicializar_archivo_asistencias(self):
        """Crea el archivo de asistencias con las columnas necesarias si no existe"""
        if not os.path.exists(ARCHIVO_ASISTENCIAS):
            wb = Workbook()
            ws = wb.active
            ws.title = "Datos"
            for idx, columna in enumerate(COLUMNAS_ASISTENCIAS, start=1):
                letra = get_column_letter(idx)
                ws[f"{letra}1"] = columna
            wb.save(ARCHIVO_ASISTENCIAS)
    
    def crear_pestana_registro(self):
        """Crea la pestaña de registro de usuarios"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Registro de Usuarios")
        
        # Encabezado
        ttk.Label(frame, text="Registro de Nuevo Usuario", style='Header.TLabel').grid(
            row=0, column=0, columnspan=2, pady=10)
        
        # Campos del formulario
        campos = [
            ("Nombre:", "nombre_entry"),
            ("Edad:", "edad_entry"),
            ("Teléfono (8 dígitos):", "telefono_entry"),
            ("Nombre del Encargado:", "encargado_entry"),
            ("Teléfono Encargado (8 dígitos):", "telefono_encargado_entry"),
            ("Correo del Encargado:", "correo_entry"),
            ("Carrera:", "carrera_entry")
        ]
        
        for i, (texto, nombre) in enumerate(campos, start=1):
            ttk.Label(frame, text=texto).grid(row=i, column=0, sticky=tk.W, padx=10, pady=5)
            entry = ttk.Entry(frame, width=30)
            entry.grid(row=i, column=1, padx=10, pady=5)
            setattr(self, nombre, entry)
        
        ttk.Button(frame, text="Registrar Usuario", command=self.registrar_usuario).grid(
            row=len(campos)+1, column=0, columnspan=2, pady=20)
        
        self.registro_message = ttk.Label(frame, text="", foreground='green')
        self.registro_message.grid(row=len(campos)+2, column=0, columnspan=2)
    
    def crear_pestana_asistencia(self):
        """Crea la pestaña para tomar asistencia con QR"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Tomar Asistencia")
        
        ttk.Label(frame, text="Tomar Asistencia con Cámara QR", style='Header.TLabel').grid(
            row=0, column=0, columnspan=2, pady=10)
        
        self.canvas = tk.Canvas(frame, width=640, height=480)
        self.canvas.grid(row=1, column=0, columnspan=2, padx=10, pady=10)
        
        # Botones
        ttk.Button(frame, text="Iniciar Cámara", command=self.iniciar_camara).grid(
            row=2, column=0, pady=10, padx=5, sticky=tk.E)
        ttk.Button(frame, text="Detener Cámara", command=self.detener_camara).grid(
            row=2, column=1, pady=10, padx=5, sticky=tk.W)
        
        self.asistencia_message = ttk.Label(frame, text="", foreground='blue')
        self.asistencia_message.grid(row=3, column=0, columnspan=2)
    
    def crear_pestana_reportes(self):
        """Crea la pestaña para generar reportes"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Reportes")
        
        ttk.Label(frame, text="Generar Reporte de Asistencias", style='Header.TLabel').grid(
            row=0, column=0, columnspan=2, pady=10)
        
        ttk.Label(frame, text="Fecha del Reporte (YYYY-MM-DD):").grid(
            row=1, column=0, sticky=tk.W, padx=10, pady=5)
        self.fecha_reporte_entry = ttk.Entry(frame, width=20)
        self.fecha_reporte_entry.grid(row=1, column=1, sticky=tk.W, padx=10, pady=5)
        self.fecha_reporte_entry.insert(0, time.strftime("%Y-%m-%d"))
        
        ttk.Button(frame, text="Generar Reporte", command=self.generar_reporte).grid(
            row=2, column=0, columnspan=2, pady=10)
        
        self.reporte_text = tk.Text(frame, height=20, width=80, wrap=tk.WORD)
        self.reporte_text.grid(row=3, column=0, columnspan=2, padx=10, pady=10)
        
        scrollbar = ttk.Scrollbar(frame, command=self.reporte_text.yview)
        scrollbar.grid(row=3, column=2, sticky='nsew')
        self.reporte_text['yscrollcommand'] = scrollbar.set
        
        ttk.Button(frame, text="Exportar Reporte", command=self.exportar_reporte).grid(
            row=4, column=0, columnspan=2, pady=10)
        
        ttk.Button(frame, text="Generar Estadísticas", command=self.generar_estadisticas_usuario).grid(
            row=5, column=0, columnspan=2, pady=10)
    
    def crear_pestana_verificacion(self):
        """Crea la pestaña para verificar asistencias"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Verificación")
        
        ttk.Label(frame, text="Verificar Asistencias por Fecha", style='Header.TLabel').grid(
            row=0, column=0, columnspan=2, pady=10)
        
        ttk.Label(frame, text="Fecha a Verificar (YYYY-MM-DD):").grid(
            row=1, column=0, sticky=tk.W, padx=10, pady=5)
        self.fecha_verificacion_entry = ttk.Entry(frame, width=20)
        self.fecha_verificacion_entry.grid(row=1, column=1, sticky=tk.W, padx=10, pady=5)
        self.fecha_verificacion_entry.insert(0, time.strftime("%Y-%m-%d"))
        
        ttk.Button(frame, text="Verificar Asistencias", command=self.verificar_asistencias).grid(
            row=2, column=0, columnspan=2, pady=10)
        
        columns = ('nombre', 'carnet', 'estado', 'correo')
        self.tree = ttk.Treeview(frame, columns=columns, show='headings', height=15)
        
        self.tree.heading('nombre', text='Nombre')
        self.tree.heading('carnet', text='Carnet')
        self.tree.heading('estado', text='Estado')
        self.tree.heading('correo', text='Correo Encargado')
        
        self.tree.column('nombre', width=200)
        self.tree.column('carnet', width=100)
        self.tree.column('estado', width=100)
        self.tree.column('correo', width=200)
        
        self.tree.grid(row=3, column=0, columnspan=2, padx=10, pady=10)
        
        scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        scrollbar.grid(row=3, column=2, sticky='ns')
        
        ttk.Button(frame, text="Enviar Notificaciones a Ausentes", command=self.enviar_notificaciones).grid(
            row=4, column=0, columnspan=2, pady=10)
    
    def validar_campos_usuario(self):
        """Valida todos los campos del formulario de registro"""

        nombre = self.nombre_entry.get().strip()
        edad = self.edad_entry.get().strip()
        telefono = self.telefono_entry.get().strip()
        encargado = self.encargado_entry.get().strip()
        telefono_encargado = self.telefono_encargado_entry.get().strip()
        correo = self.correo_entry.get().strip()
        carrera = self.carrera_entry.get().strip()
        
        campos = {
            "Nombre": nombre,
            "Edad": edad,
            "Teléfono": telefono,
            "Nombre del encargado": encargado,
            "Teléfono del encargado": telefono_encargado,
            "Correo del encargado": correo,
            "Carrera": carrera
        }
        
        for campo, valor in campos.items():
            if not valor:
                messagebox.showerror("Error", f"El campo {campo} es obligatorio")
                return False
        
        if not edad.isdigit() or int(edad) <= 0:
            messagebox.showerror("Error", "La edad debe ser un número positivo")
            return False
        
        if not telefono.isdigit() or len(telefono) != 8:
            messagebox.showerror("Error", "El teléfono debe tener exactamente 8 dígitos")
            return False
            
        if not telefono_encargado.isdigit() or len(telefono_encargado) != 8:
            messagebox.showerror("Error", "El teléfono del encargado debe tener exactamente 8 dígitos")
            return False
        
        if "@" not in correo or "." not in correo:
            messagebox.showerror("Error", "Ingrese un correo electrónico válido")
            return False
        
        if any(caracter.isdigit() for caracter in nombre):
            messagebox.showerror("Error", "El nombre no puede contener números")
            return False
            
        if any(caracter.isdigit() for caracter in encargado):
            messagebox.showerror("Error", "El nombre del encargado no puede contener números")
            return False
        
        return True

    def registrar_usuario(self):
        """Registra un nuevo usuario en el sistema"""
        try:
            if not self.validar_campos_usuario():
                return
                
            nombre = self.nombre_entry.get().strip()
            edad = self.edad_entry.get().strip()
            telefono = self.telefono_entry.get().strip()
            encargado = self.encargado_entry.get().strip()
            telefono_encargado = self.telefono_encargado_entry.get().strip()
            correo = self.correo_entry.get().strip()
            carrera = self.carrera_entry.get().strip()
            
            carnet = self.obtener_siguiente_carnet(NOMBRE_CARNET)
            
            confirmacion = messagebox.askyesno(
                "Confirmar registro",
                f"¿Desea registrar al usuario?\n\n"
                f"Nombre: {nombre}\n"
                f"Carnet: {carnet}\n"
                f"Edad: {edad}\n"
                f"Teléfono: {telefono}\n"
                f"Carrera: {carrera}\n"
                f"Encargado: {encargado}\n"
                f"Teléfono encargado: {telefono_encargado}\n"
                f"Correo encargado: {correo}"
            )
            
            if not confirmacion:
                return
                
            self.agregar_datos_usuario(nombre, carnet, edad, telefono, encargado, telefono_encargado, correo, carrera)
            
            self.registro_message.config(
                text=f"Usuario {nombre} registrado exitosamente con carnet {carnet}",
                foreground="green"
            )
            
            for entry in [self.nombre_entry, self.edad_entry, self.telefono_entry, 
                        self.encargado_entry, self.telefono_encargado_entry, 
                        self.correo_entry, self.carrera_entry]:
                entry.delete(0, tk.END)
            
            self.enviar_correo_registro(correo, nombre, carnet)
            
        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al registrar el usuario: {str(e)}")
        
    def iniciar_camara(self):
        """Inicia la cámara para leer códigos QR"""
        if self.running:
            return
            
        self.running = True
        self.cap = cv2.VideoCapture(0)
        self.asistencia_message.config(text="Escanea el código QR... Presiona 'Detener Cámara' para salir.")
        self.leer_qr()
    
    def detener_camara(self):
        """Detiene la cámara"""
        self.running = False
        if self.cap:
            self.cap.release()
            self.cap = None
        self.asistencia_message.config(text="Cámara detenida")

    def verificar_carnet(self,carnet: str) -> bool:
        """Verifica si el carnet existe en la base de datos"""
        try:
            wb = load_workbook(NOMBRE_CARNET)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                if row[1] == carnet:
                    print("Carnet Correcto")
                    return True
            print("Carnet Inválido")
            return False
        except Exception as e:
            print(f"Error al verificar carnet: {e}")
            return False
        
    
    def registrar_asistencia(self,nombre: str, carnet: str) -> bool:
        """Registra la asistencia en el archivo Excel si no está ya registrada"""
        try:
            wb = load_workbook(ARCHIVO_ASISTENCIAS)
            ws = wb.active
            
            fecha_actual = time.strftime("%Y-%m-%d")
            hora_actual = time.strftime("%H:%M:%S")
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[1] == carnet and row[2] == fecha_actual:
                    print("La asistencia ya fue registrada hoy.")
                    return False
            
            nueva_fila = [nombre, carnet, fecha_actual, hora_actual]
            ws.append(nueva_fila)
            wb.save(ARCHIVO_ASISTENCIAS)
            # self.actualizar_estadisticas(carnet, nombre, "Carrera", asistio=True)  
            print("Asistencia registrada exitosamente.")
            return True
        except Exception as e:
            print(f"Error al registrar asistencia: {e}")
            return False
        
    def procesar_datos_qr(self,data: str) -> tuple:
        """Extrae y valida los datos del QR"""
        try:
            if "Carnet:" not in data or "Nombre:" not in data:
                print("Formato incorrecto: faltan 'Carnet:' o 'Nombre:' en la cadena.")
                return None, None
            
            carnet = data.split("Carnet:")[1].split("Nombre:")[0].strip()
            nombre = data.split("Nombre:")[1].strip()
            return nombre, carnet
        except Exception as e:
            print(f"Error al procesar QR: {e}")
            return None, None
        
    def leer_qr(self):
        """Lee códigos QR de la cámara"""

        if not self.running or not self.cap:
            return
            
        ret, frame = self.cap.read()
        if ret:
            frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            img = Image.fromarray(frame)
            imgtk = ImageTk.PhotoImage(image=img)            
            self.canvas.imgtk = imgtk
            self.canvas.create_image(0, 0, anchor=tk.NW, image=imgtk)
            
            codigos = decode(frame)
            if codigos:
                for codigo in codigos:
                    try:
                        data = codigo.data.decode('utf-8')
                        nombre, carnet = self.procesar_datos_qr(data)
                        
                        if nombre and carnet and self.verificar_carnet(carnet):
                            if self.registrar_asistencia(nombre, carnet):
                                self.asistencia_message.config(
                                    text=f"Asistencia registrada: {nombre} ({carnet})", 
                                    foreground='green')
                                self.mqtt_cliente.publish(TOPIC, "0") 
                                self.mqtt_cliente.publish(TOPIC_PANTALLA,f"{carnet}")
                            else:
                                self.asistencia_message.config(
                                    text="Asistencias ya registradas hoy", 
                                    foreground='orange')
                                self.mqtt_cliente.publish(TOPIC, "1")
                                self.mqtt_cliente.publish(TOPIC_PANTALLA, "Asistencia Registrada")  
                        else:
                            self.asistencia_message.config(
                                text="Código QR no válido", 
                                foreground='red')
                            self.mqtt_cliente.publish(TOPIC_PANTALLA,"Codigo QR no Valido")
                            self.mqtt_cliente.publish(TOPIC, "1")  
                            
                    except Exception as e:
                        print(f"Error al procesar código QR: {e}")
                        self.mqtt_cliente.publish(TOPIC, "1")  
        
        if self.running:
            self.root.after(10, self.leer_qr)

    def cerrar(self):
        """Limpieza al cerrar"""
        self.mqtt_cliente.loop_stop()
        self.mqtt_cliente.disconnect()

    def obtener_lista_usuarios(self):
        """Obtiene la lista completa de usuarios registrados"""
        try:
            wb = load_workbook(ARCHIVO_USUARIOS)
            ws = wb.active
            usuarios = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Si tiene nombre
                    usuarios.append({
                        'Nombre': row[0],
                        'Carnet': row[1],
                        'Correo Encargado': row[6] if len(row) > 6 else 'No especificado',
                        'Carrera': row[7] if len(row) > 7 else 'No especificado'
                    })
            return usuarios
        except Exception as e:
            print(f"Error al obtener usuarios: {e}")
            return []

    def obtener_asistencias_por_fecha(self,fecha):
        """Obtiene las asistencias de una fecha específica"""
        try:
            wb = load_workbook(ARCHIVO_ASISTENCIAS)
            ws = wb.active
            asistencias = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[2] == fecha:
                    asistencias.append(row[1])  # Guardamos solo el carnet
            return asistencias
        except Exception as e:
            print(f"Error al obtener asistencias: {e}")
            return []
    
    def generar_reporte(self):
        """Genera un reporte de asistencias para la fecha especificada en pantalla y en archivo Excel"""
        try:
            fecha = self.fecha_reporte_entry.get().strip()
            
            # Validar formato de fecha
            if len(fecha) != 10 or fecha[4] != '-' or fecha[7] != '-':
                messagebox.showerror("Error", "Formato de fecha inválido. Use YYYY-MM-DD")
                return
                
            # Obtener datos
            usuarios = self.obtener_lista_usuarios()
            asistencias = self.obtener_asistencias_por_fecha(fecha)
            
            presentes = []
            ausentes = []
            
            for usuario in usuarios:
                if usuario['Carnet'] in asistencias:
                    presentes.append(usuario)
                else:
                    ausentes.append(usuario)
            
            # Mostrar en el área de texto
            self.reporte_text.delete(1.0, tk.END)
            self.reporte_text.insert(tk.END, f"=== Reporte de Asistencias ===\n")
            self.reporte_text.insert(tk.END, f"Fecha: {fecha}\n\n")
            
            self.reporte_text.insert(tk.END, f"Presentes ({len(presentes)}):\n")
            for usuario in presentes:
                self.reporte_text.insert(tk.END, f"- {usuario['Nombre']} ({usuario['Carnet']}) - {usuario['Carrera']}\n")
            
            self.reporte_text.insert(tk.END, f"\nAusentes ({len(ausentes)}):\n")
            for usuario in ausentes:
                self.reporte_text.insert(tk.END, f"- {usuario['Nombre']} ({usuario['Carnet']}) - {usuario['Carrera']}\n")
            
            # Crear archivo Excel
            self.generar_excel_reporte(fecha, presentes, ausentes)
            
            # self.registrar_ausencias(fecha) 
            messagebox.showinfo("Éxito", "Reporte generado correctamente y guardado en Excel")
            
        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al generar el reporte: {str(e)}")

    def generar_excel_reporte(self, fecha, presentes, ausentes):
        """Genera un archivo Excel con el reporte de asistencias"""
        try:
            # Crear directorio de reportes si no existe
            reportes_dir = os.path.join(DIR_ASISTENCIAS, "Reportes_Excel")
            os.makedirs(reportes_dir, exist_ok=True)
            
            # Crear nombre de archivo con fecha
            nombre_archivo = os.path.join(reportes_dir, f"reporte_asistencia_{fecha}.xlsx")
            
            # Crear libro de Excel
            wb = Workbook()
            
            # Hoja para presentes
            ws_presentes = wb.active
            ws_presentes.title = "Presentes"
            ws_presentes.append(["Nombre", "Carnet", "Carrera"])
            
            for usuario in presentes:
                ws_presentes.append([usuario['Nombre'], usuario['Carnet'], usuario['Carrera']])
            
            # Hoja para ausentes
            ws_ausentes = wb.create_sheet(title="Ausentes")
            ws_ausentes.append(["Nombre", "Carnet", "Carrera", "Correo Encargado"])
            
            for usuario in ausentes:
                ws_ausentes.append([
                    usuario['Nombre'], 
                    usuario['Carnet'], 
                    usuario['Carrera'],
                    usuario.get('Correo Encargado', 'No especificado')
                ])
            
            # Hoja de resumen
            ws_resumen = wb.create_sheet(title="Resumen")
            ws_resumen.append(["Tipo", "Cantidad"])
            ws_resumen.append(["Presentes", len(presentes)])
            ws_resumen.append(["Ausentes", len(ausentes)])
            ws_resumen.append(["Total", len(presentes) + len(ausentes)])
            
            # Ajustar anchos de columnas
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    ws.column_dimensions[column].width = adjusted_width
            
            # Guardar archivo
            wb.save(nombre_archivo)
            
            print(f"Reporte Excel guardado en: {nombre_archivo}")
            
        except Exception as e:
            print(f"Error al generar Excel: {e}")
            raise
        
    def exportar_reporte(self):
        """Exporta el reporte a un archivo de texto"""
        try:
            reporte = self.reporte_text.get(1.0, tk.END)
            if not reporte.strip():
                messagebox.showwarning("Advertencia", "No hay reporte para exportar")
                return
                
            fecha = self.fecha_reporte_entry.get().strip()
            default_filename = f"reporte_asistencia_{fecha}.txt"
            
            filepath = filedialog.asksaveasfilename(
                defaultextension=".txt",
                initialfile=default_filename,
                filetypes=[("Archivos de texto", "*.txt"), ("Todos los archivos", "*.*")])
                
            if filepath:
                with open(filepath, 'w') as f:
                    f.write(reporte)
                messagebox.showinfo("Éxito", f"Reporte exportado a:\n{filepath}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al exportar el reporte: {str(e)}")
    
    def verificar_asistencias(self):
        """Verifica las asistencias para una fecha específica"""
        try:
            fecha = self.fecha_verificacion_entry.get().strip()
            
            if len(fecha) != 10 or fecha[4] != '-' or fecha[7] != '-':
                messagebox.showerror("Error", "Formato de fecha inválido. Use YYYY-MM-DD")
                return
                
            for item in self.tree.get_children():
                self.tree.delete(item)
                
            wb_usuarios = load_workbook(ARCHIVO_USUARIOS)
            ws_usuarios = wb_usuarios.active
            todos_usuarios = [
                (row[0], str(row[1]), row[6] ) # nombre, carnet, correo encargado
                for row in ws_usuarios.iter_rows(min_row=2, values_only=True)
            ]
            
            wb_asistencias = load_workbook(ARCHIVO_ASISTENCIAS)
            ws_asistencias = wb_asistencias.active
            asistencias_fecha = [
                (row[0], str(row[1])) 
                for row in ws_asistencias.iter_rows(min_row=2, values_only=True)
                if row[2] == fecha
            ]
            carnets_asistieron = set(c[1] for c in asistencias_fecha)
            
            for nombre, carnet, correo_tutor in todos_usuarios:
                if carnet in carnets_asistieron:
                    estado = "PRESENTE"
                    tags = ('presente',)
                else:
                    estado = "AUSENTE"
                    tags = ('ausente',)
                
                self.tree.insert('', tk.END, values=(nombre, carnet, estado, correo_tutor), tags=tags)
            
            self.tree.tag_configure('presente', foreground='green')
            self.tree.tag_configure('ausente', foreground='red')
            
            messagebox.showinfo("Éxito", "Verificación de asistencias completada")
            
        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al verificar asistencias: {str(e)}")

    def enviar_correo(self,destinatario,fecha,nombre,carnet):
        """Envía un correo electrónico con el asunto y cuerpo especificados"""

        cuerpo = f"""
        Estimado/a Encargado,

        Le saludamos cordialmente del sistema de control de asistencias. Le informamos que el estudiante asociado a su contacto no ha registrado su asistencia el día de hoy.

        Es importante que esté al tanto de esta situación y, si considera necesario, se comunique con el estudiante para conocer el motivo de su ausencia.

        Detalles:
        📅 Fecha de inasistencia: {fecha}
        🆔 Carnet: {carnet}
        👤 Estudiante: {nombre}

        Si tiene alguna duda o necesita más información, no dude en contactarnos.

        Atentamente,

        Sistema de Gestión de Asistencias
        """

        em = EmailMessage()
        em["From"] = email_sender
        em["To"] = destinatario
        em["Subject"] = "Asistencia"
        em.set_content(cuerpo)
        context = ssl.create_default_context()
        try:
            with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as smtp:
                smtp.login(email_sender, password)
                smtp.sendmail(email_sender, destinatario, em.as_string())
                print("Correo enviado exitosamente.")
        except Exception as e:
            print(f"Error al enviar correo: {e}")
        
    def enviar_notificaciones(self):
        """Envía notificaciones por correo a los ausentes"""
        try:
            # Obtener ausentes del treeview
            ausentes = []
            for item in self.tree.get_children():
                values = self.tree.item(item, 'values')
                if values and values[2] == "AUSENTE":
                    ausentes.append({
                        'nombre': values[0],
                        'carnet': values[1],
                        'correo': values[3]
                    })
            
            if not ausentes:
                messagebox.showinfo("Información", "No hay ausentes para notificar")
                return
                
            # Confirmar envío
            confirmar = messagebox.askyesno(
                "Confirmar", 
                f"¿Desea enviar notificaciones a {len(ausentes)} ausentes?")
                
            if not confirmar:
                return
                
            # Enviar correos
            for ausente in ausentes:
                if ausente['correo'] and ausente['correo'].strip():
                    try:
                        self.enviar_correo(ausente['correo'].strip(),time.strftime("%Y-%m-%d"), ausente['nombre'], ausente['carnet']    )
                        print(f"Notificación enviada a {ausente['nombre']} ({ausente['correo']})")
                    except Exception as e:
                        print(f"Error al enviar a {ausente['correo']}: {str(e)}")
            
            messagebox.showinfo("Éxito", "Proceso de notificación completado")
            
        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al enviar notificaciones: {str(e)}")

    def enviar_correo_registro(self, destinatario, nombre, carnet):
        """Envía un correo electrónico al encargado al registrar un nuevo usuario"""
        cuerpo = f"""
        Estimado/a Encargado,

        Le informamos que se ha registrado un nuevo usuario en el sistema de control de asistencias.

        Detalles:
        🆔 Carnet: {carnet}
        👤 Nombre: {nombre}

        Si tiene alguna duda o necesita más información, no dude en contactarnos.

        Atentamente,

        Sistema de Gestión de Asistencias
        """

        em = EmailMessage()
        em["From"] = email_sender
        em["To"] = destinatario
        em["Subject"] = "Registro de Nuevo Usuario"
        em.set_content(cuerpo)
        
        context = ssl.create_default_context()
        try:
            with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as smtp:
                smtp.login(email_sender, password)
                smtp.sendmail(email_sender, destinatario, em.as_string())
                print("Correo de registro enviado exitosamente.")
        except Exception as e:
            print(f"Error al enviar correo de registro: {e}")


    def generar_estadisticas_usuario(self):
        """Genera un archivo Excel con estadísticas de asistencia por usuario"""
        try:
            # Cargar usuarios y asistencias
            wb_usuarios = load_workbook(ARCHIVO_USUARIOS)
            ws_usuarios = wb_usuarios.active

            wb_asistencias = load_workbook(ARCHIVO_ASISTENCIAS)
            ws_asistencias = wb_asistencias.active

            # Organizar asistencias por carnet
            asistencias_por_carnet = {}
            for row in ws_asistencias.iter_rows(min_row=2, values_only=True):
                nombre, carnet, fecha, hora = row
                if carnet not in asistencias_por_carnet:
                    asistencias_por_carnet[carnet] = []
                asistencias_por_carnet[carnet].append(fecha)

            # Crear directorio si no existe
            reportes_dir = os.path.join(DIR_ASISTENCIAS, "Reportes_Estadisticas")
            os.makedirs(reportes_dir, exist_ok=True)

            # Nombre del archivo
            nombre_archivo = os.path.join(reportes_dir, "estadisticas_asistencia.xlsx")

            # Crear libro de Excel
            wb = Workbook()
            ws_resumen = wb.active
            ws_resumen.title = "Resumen General"

            # Encabezados
            encabezados = [
                "Nombre", "Carnet", "Total Registros", 
                "Asistencias", "Ausencias", "Porcentaje Asistencia"
            ]
            for col_num, data in enumerate(encabezados, 1):
                ws_resumen.cell(row=1, column=col_num, value=data)

            fila = 2
            detalles = []

            for row in ws_usuarios.iter_rows(min_row=2, values_only=True):
                nombre = row[0]
                carnet = str(row[1])

                todas_fechas = sorted(set([fecha for _, _, fecha, _ in ws_asistencias.iter_rows(min_row=2, values_only=True)]))
                total_dias = len(todas_fechas)

                asistencias_usuario = asistencias_por_carnet.get(carnet, [])
                asistencias_count = len(asistencias_usuario)
                ausencias_count = total_dias - asistencias_count
                porcentaje = (asistencias_count / total_dias * 100) if total_dias > 0 else 0

                # Agregar a resumen general
                ws_resumen.append([
                    nombre, carnet, total_dias,
                    asistencias_count, ausencias_count, f"{porcentaje:.2f}%"
                ])

                # Detalles por usuario para otra hoja
                fechas_presentes = set(asistencias_usuario)
                fechas_ausentes = [f for f in todas_fechas if f not in fechas_presentes]
                detalles.append({
                    "nombre": nombre,
                    "carnet": carnet,
                    "fechas_ausentes": fechas_ausentes
                })

            # Hoja de detalles
            ws_detalles = wb.create_sheet(title="Detalles Ausencias")
            ws_detalles.append(["Nombre", "Carnet", "Fechas Ausentes"])

            for detalle in detalles:
                fechas_str = ", ".join(detalle["fechas_ausentes"]) if detalle["fechas_ausentes"] else "Ninguna"
                ws_detalles.append([detalle["nombre"], detalle["carnet"], fechas_str])

            # Ajustar anchos de columna
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    ws.column_dimensions[column].width = adjusted_width

            # Guardar archivo
            wb.save(nombre_archivo)
            messagebox.showinfo("Éxito", f"Estadísticas generadas correctamente en:\n{nombre_archivo}")

        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al generar las estadísticas: {str(e)}")
            

# Ejecutar la aplicación
if __name__ == "__main__":
    root = tk.Tk()
    app = SistemaAsistenciasApp(root)
    root.mainloop()