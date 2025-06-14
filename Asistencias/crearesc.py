from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os

# Nombre del archivo
archivo = 'datos.xlsx'

# Nombres de las columnas (puedes modificarlos)
columnas = ["Nombre", "Edad", "Correo","Telefono","Encargado","Telefono Encargado","Correo Encargado"]

# Si el archivo no existe, lo crea y escribe los encabezados
if not os.path.exists(archivo):
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"

    for idx, columna in enumerate(columnas, start=1):
        letra = get_column_letter(idx)
        ws[f"{letra}1"] = columna

    wb.save(archivo)

# Función para agregar una fila de datos
def agregar_dato(nombre, edad, correo,telefono,encargado,telefono_encargado,correo_encargado):
    if not os.path.exists(archivo):
        print("El archivo no existe. Creando uno nuevo...")
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos"
        for idx, columna in enumerate(columnas, start=1):
            letra = get_column_letter(idx)
            ws[f"{letra}1"] = columna
        wb.save(archivo)
    else:
        print("El archivo ya existe. Agregando datos...")
    wb = load_workbook(archivo)
    ws = wb.active

    nueva_fila = [nombre, edad, correo, telefono, encargado, telefono_encargado,correo_encargado]
    # Verificar si la fila ya existe
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == nombre and row[1] == edad and row[2] == correo:
            print("El dato ya existe en la hoja de cálculo.")
            return
    # Si no existe, agregar la nueva fila
    max_row = ws.max_row        
    ws.append(nueva_fila)
    wb.save(archivo)
    print("Dato agregado correctamente.")

# Ejemplo de uso
agregar_dato("Juan Pérez", 30, "juanperez@example.com", "12345678", "Ana Pérez", "87654321")
agregar_dato("María López", 25, "marialopez@example.com", "23456789", "Carlos López", "98765432")
